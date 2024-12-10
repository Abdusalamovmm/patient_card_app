from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import openpyxl.utils as utils
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import Workbook
from io import BytesIO
import os


def generate_excel_report(patients):
    """Генерация отчета в формате Excel с форматированием."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Пациенты и болезни"
    headers = ["Фотография", "ФИО", "Дата рождения", "Телефон", "Адрес", "Болезни", "Дата заболевания", "Дата выздоровления",
               "Комментарий"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        column_letter = utils.get_column_letter(col_num)
        if header == "Фотография":
            sheet.column_dimensions[column_letter].width = 12
        elif header == "ФИО":
            sheet.column_dimensions[column_letter].width = 25
        elif header == "Адрес":
            sheet.column_dimensions[column_letter].width = 25
        elif header == "Комментарий":
            sheet.column_dimensions[column_letter].width = 40
        else:
            sheet.column_dimensions[column_letter].width = 15
        sheet.row_dimensions[1].height = 30
    row_num = 2
    for patient in patients:
        patient_diseases = patient.get('diseases', [])
        photo_added = False
        if patient_diseases:
            for disease in patient_diseases:
                if not photo_added and patient.get('photo'):
                    try:
                        img = Image(patient['photo'])
                        img.width = 64
                        img.height = 64
                        sheet.add_image(img, f'A{row_num}')
                        photo_added = True
                    except Exception as e:
                        print(f"Error adding image: {e}")

                sheet.cell(row=row_num, column=2).value = patient['name']
                sheet.cell(row=row_num, column=2).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=3).value = patient['birth_date']
                sheet.cell(row=row_num, column=3).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=4).value = patient['phone']
                sheet.cell(row=row_num, column=4).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=5).value = patient['address']
                sheet.cell(row=row_num, column=5).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=6).value = disease['disease_name']
                sheet.cell(row=row_num, column=6).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=7).value = disease['start_date']
                sheet.cell(row=row_num, column=7).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=8).value = disease['recovery_date']
                sheet.cell(row=row_num, column=8).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
                sheet.cell(row=row_num, column=9).value = disease['doctor_comment']
                sheet.cell(row=row_num, column=9).alignment = Alignment(horizontal='center', vertical='center',
                                                                        wrap_text=True)
                sheet.row_dimensions[row_num].height = 60
                row_num += 1
        else:
            if not photo_added and patient.get('photo'):
                try:
                    img = Image(patient['photo'])
                    img.width = 64
                    img.height = 64
                    sheet.add_image(img, f'A{row_num}')
                    photo_added = True
                except Exception as e:
                    print(f"Error adding image: {e}")
            sheet.cell(row=row_num, column=2).value = patient['name']
            sheet.cell(row=row_num, column=3).value = patient['birth_date']
            sheet.cell(row=row_num, column=4).value = patient['phone']
            sheet.cell(row=row_num, column=5).value = patient['address']
            sheet.row_dimensions[row_num].height = 60
            row_num += 1
    return workbook


def save_report_to_excel(workbook):
    """Сохранение отчета в Excel файл."""
    try:
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filepath:
            return
        workbook.save(filepath)
        messagebox.showinfo("Успех", f"Отчет сохранен в файл: {filepath}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при сохранении отчета: {e}")
